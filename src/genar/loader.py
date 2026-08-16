from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def load_rows(path: str | Path) -> list[dict[str, Any]]:
    """Load the challenge flat file from CSV or XLSX.

    XLSX support intentionally uses openpyxl only at the ingestion boundary. The rest
    of the pipeline receives plain Python dictionaries and is format-independent.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [{k: _clean(v) for k, v in row.items()} for row in csv.DictReader(f)]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX input requires openpyxl. Run: pip install -r requirements.txt") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        headers = [str(v).strip() for v in next(iterator)]
        out: list[dict[str, Any]] = []
        for values in iterator:
            out.append({h: _clean(v) for h, v in zip(headers, values)})
        wb.close()
        return out
    raise ValueError(f"Unsupported data format: {path.suffix}. Use .csv or .xlsx")


def _as_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def deduplicate_latest(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Keep the highest safetyreportversion for each safetyreportid.

    Follow-up versions are separate rows in the source. Case-level metrics must not
    count those as separate patients/cases.
    """
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    row_count = 0
    for row in rows:
        row_count += 1
        case_id = str(row.get("safetyreportid") or "").strip()
        if not case_id:
            raise ValueError("Encountered a row without safetyreportid")
        grouped[case_id].append(row)

    latest: list[dict[str, Any]] = []
    duplicate_case_ids: list[str] = []
    for case_id, versions in grouped.items():
        if len(versions) > 1:
            duplicate_case_ids.append(case_id)
        latest.append(
            max(
                versions,
                key=lambda r: (
                    _as_int(r.get("safetyreportversion")),
                    sum(v not in (None, "") for v in r.values()),
                ),
            )
        )

    latest.sort(key=lambda r: (parse_yyyymmdd(r.get("receivedate")) or date.min, str(r.get("safetyreportid"))))
    audit = {
        "source_rows": row_count,
        "unique_cases": len(latest),
        "followup_rows_removed": row_count - len(latest),
        "case_ids_with_multiple_versions": len(duplicate_case_ids),
    }
    return latest, audit


def parse_yyyymmdd(value: Any) -> date | None:
    if value in (None, ""):
        return None
    try:
        digits = str(int(float(value)))
    except (TypeError, ValueError):
        digits = str(value).strip()
    if len(digits) != 8 or not digits.isdigit():
        return None
    try:
        return datetime.strptime(digits, "%Y%m%d").date()
    except ValueError:
        return None


def excel_serial_to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    except (TypeError, ValueError, OverflowError):
        return None


def split_multi(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def parse_reactions(row: dict[str, Any]) -> list[str]:
    """Parse flattened MedDRA PT arrays while preserving PTs containing commas.

    The challenge file flattens arrays with commas. A few valid PTs themselves contain
    commas (for example 'Hallucination, visual'). The parallel MedDRA-version field
    tells us the expected number of reactions. When naive splitting produces extra
    tokens, lowercase modifier fragments are merged back into the preceding PT.
    """
    raw = row.get("patient_reaction_reactionmeddrapt")
    if raw in (None, ""):
        return []
    tokens = [t.strip() for t in str(raw).split(",") if t.strip()]
    versions = split_multi(row.get("patient_reaction_reactionmeddraversionpt"))
    expected = len(versions) or None

    if expected and len(tokens) > expected:
        i = 1
        while len(tokens) > expected and i < len(tokens):
            token = tokens[i]
            if token and token[0].islower():
                tokens[i - 1] = f"{tokens[i - 1]}, {token}"
                tokens.pop(i)
            else:
                i += 1

    if expected and len(tokens) != expected:
        # Preserve data but make the ambiguity observable upstream.
        row.setdefault("_reaction_parse_warning", True)
    return tokens
